"""
Discord Members → Excel Exporter (mit GUI + Supabase Minecraft-Namen)
======================================================================
GUI zum Auswählen der zu exportierenden Daten.
Benötigt: pip install discord.py openpyxl python-dotenv supabase
"""
import os
import math
import asyncio
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime, timezone

import discord
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

MINECRAFT_COLORS = [
    {"name": "Weiß",      "hex": "F9FFFE", "rgb": (249, 255, 254)},
    {"name": "Hellgrau",  "hex": "D1D1D1", "rgb": (209, 209, 209)},
    {"name": "Grau",      "hex": "3C3C3C", "rgb": (60,  60,  60)},
    {"name": "Schwarz",   "hex": "1D1D21", "rgb": (29,  29,  33)},
    {"name": "Braun",     "hex": "835432", "rgb": (131, 84,  50)},
    {"name": "Rot",       "hex": "B02E26", "rgb": (176, 46,  38)},
    {"name": "Orange",    "hex": "F9801D", "rgb": (249, 128, 29)},
    {"name": "Gelb",      "hex": "FED83D", "rgb": (254, 216, 61)},
    {"name": "Hellgrün",  "hex": "80C71F", "rgb": (128, 199, 31)},
    {"name": "Grün",      "hex": "5E7C16", "rgb": (94,  124, 22)},
    {"name": "Cyan",      "hex": "169C9C", "rgb": (22,  156, 156)},
    {"name": "Hellblau",  "hex": "3AB3DA", "rgb": (58,  179, 218)},
    {"name": "Blau",      "hex": "3C44AA", "rgb": (60,  68,  170)},
    {"name": "Lila",      "hex": "8932B8", "rgb": (137, 50,  184)},
    {"name": "Magenta",   "hex": "C74EBD", "rgb": (199, 78,  189)},
    {"name": "Rosa",      "hex": "F38BAA", "rgb": (243, 139, 170)},
]

def color_distance(r1, g1, b1, r2, g2, b2):
    return math.sqrt((r1-r2)**2 + (g1-g2)**2 + (b1-b2)**2)

def closest_minecraft_color(val):
    if val == 0:
        r, g, b = 153, 153, 153
    else:
        r = (val >> 16) & 0xFF
        g = (val >> 8)  & 0xFF
        b =  val        & 0xFF
    best = min(MINECRAFT_COLORS, key=lambda c: color_distance(r, g, b, *c["rgb"]))
    return best["name"], best["hex"]

def get_highest_role(member):
    roles = [r for r in member.roles if r.name != "@everyone"]
    return max(roles, key=lambda r: r.position) if roles else None

def fmt_dt(dt):
    return dt.strftime("%d.%m.%Y %H:%M") if dt else ""

def days_since(dt):
    if not dt:
        return ""
    return (datetime.now(timezone.utc) - dt).days

def txt_on_bg(hex6):
    r, g, b = int(hex6[0:2],16), int(hex6[2:4],16), int(hex6[4:6],16)
    return "000000" if (0.299*r + 0.587*g + 0.114*b) > 128 else "FFFFFF"

def color_cell(cell, hex6, bold=False):
    cell.fill      = PatternFill("solid", start_color=hex6)
    cell.font      = Font(name="Arial", size=10, bold=bold, color=txt_on_bg(hex6))
    cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_header(ws, headers, widths, bg="23272A"):
    fill = PatternFill("solid", start_color=bg)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    ctr  = Alignment(horizontal="center", vertical="center")
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font; c.fill = fill; c.alignment = ctr
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

LEFT   = Alignment(horizontal="left",   vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

class Exporter(discord.Client):
    def __init__(self, token, guild_id, output_path, options, log_callback=None):
        intents = discord.Intents.default()
        intents.members = True
        intents.presences = True
        super().__init__(intents=intents)
        self._token = token
        self.guild_id = guild_id
        self.output_path = output_path
        self.options = options
        self.log = log_callback or print
        self.supabase: Client = None
        if SUPABASE_URL and SUPABASE_KEY:
            try:
                self.supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
                self.log("✅ Supabase-Client initialisiert")
            except Exception as e:
                self.log(f"⚠️ Supabase-Fehler: {e}")

    async def fetch_minecraft_names(self, user_ids):
        if not self.supabase or not self.options.get("include_minecraft_name", False):
            return {}
        names = {}
        try:
            response = (
                self.supabase.table("applications")
                .select("creator_id, minecraft_name, created_at")
                .in_("creator_id", user_ids)
                .order("created_at", desc=True)
                .execute()
            )
            if response.data:
                seen = set()
                for row in response.data:
                    uid = row["creator_id"]
                    if uid not in seen:
                        seen.add(uid)
                        names[uid] = row["minecraft_name"] or ""
        except Exception as e:
            self.log(f"⚠️ Fehler beim Abrufen der Minecraft-Namen: {e}")
        return names

    async def on_ready(self):
        guild = self.get_guild(self.guild_id)
        if guild is None:
            self.log(f"❌ Server {self.guild_id} nicht gefunden.")
            await self.close()
            return

        self.log(f"✅ Verbunden – {guild.name} ({guild.member_count} Mitglieder)")
        self.log("📋 Lade Mitglieder...")
        try:
            members = [m async for m in guild.fetch_members(limit=None)]
        except Exception as e:
            self.log(f"❌ Fehler: {e}")
            await self.close()
            return

        total = len(members)
        self.log(f"   {total} Mitglieder geladen. Erstelle Excel...")

        minecraft_names = {}
        if self.options.get("include_minecraft_name", False):
            self.log("📋 Lade Minecraft-Namen aus Supabase...")
            user_ids = [str(m.id) for m in members]
            minecraft_names = await self.fetch_minecraft_names(user_ids)
            self.log(f"   {len(minecraft_names)} Minecraft-Namen gefunden.")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        bots = sum(1 for m in members if m.bot)
        humans = total - bots

        # Mitglieder
        if self.options.get("export_members", True):
            ws1 = wb.create_sheet("Mitglieder")
            all_headers = [
                "User-ID", "Username", "Anzeigename (Server)", "Globaler Name",
                "Discriminator", "Bot?", "System-Account?", "Online-Status",
                "Account erstellt", "Server beigetreten", "Tage Mitglied",
                "Account-Alter (Tage)", "Anzahl Rollen", "Höchste Rolle",
                "Rollenfarbe (HEX)", "Minecraft-Farbe", "MC-Farbe (HEX)",
                "Alle Rollen", "Administrator?", "Nachrichten verwalten?",
                "Kick?", "Ban?", "Moderator (Timeout)?", "Kanäle verwalten?",
                "Rollen verwalten?", "Server verwalten?", "Webhooks verwalten?",
                "Avatar-URL", "Server-Booster?", "Boosted seit", "Timeout bis",
                "Minecraft-Name", "Gebaut"
            ]
            all_widths = [
                20, 24, 24, 24, 14, 7, 13, 22, 18, 20, 14, 18, 14, 24,
                18, 18, 14, 70, 14, 22, 8, 8, 22, 20, 18, 18, 20, 70,
                15, 18, 18, 20, 10
            ]
            selected_indices = self.options.get("member_columns", list(range(len(all_headers))))
            headers = [all_headers[i] for i in selected_indices]
            widths  = [all_widths[i] for i in selected_indices]
            apply_header(ws1, headers, widths)
            ws1.freeze_panes = "A2"
            ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

            STATUS_MAP = {
                discord.Status.online: "🟢 Online",
                discord.Status.idle: "🌙 Abwesend",
                discord.Status.dnd: "⛔ Bitte nicht stören",
                discord.Status.offline: "⚫ Offline",
            }

            for idx, member in enumerate(members, 2):
                hr = get_highest_role(member)
                role_name = hr.name if hr else "Kein Rang"
                role_color = hr.color.value if hr else 0
                mc_name, mc_hex = closest_minecraft_color(role_color)
                no_everyone = [r for r in member.roles if r.name != "@everyone"]
                all_roles = ", ".join(r.name for r in sorted(no_everyone, key=lambda r: -r.position))
                p = member.guild_permissions
                status_str = STATUS_MAP.get(member.status, str(member.status))
                mc_username = minecraft_names.get(str(member.id), "")

                all_values = [
                    str(member.id), str(member), member.display_name,
                    getattr(member, "global_name", "") or "",
                    member.discriminator if member.discriminator != "0" else "",
                    "Ja" if member.bot else "Nein",
                    "Ja" if member.system else "Nein", status_str,
                    fmt_dt(member.created_at), fmt_dt(member.joined_at),
                    days_since(member.joined_at), days_since(member.created_at),
                    len(no_everyone), role_name,
                    f"#{role_color:06X}" if role_color else "#000000",
                    mc_name, mc_hex, all_roles,
                    "Ja" if p.administrator else "Nein",
                    "Ja" if p.manage_messages else "Nein",
                    "Ja" if p.kick_members else "Nein",
                    "Ja" if p.ban_members else "Nein",
                    "Ja" if p.moderate_members else "Nein",
                    "Ja" if p.manage_channels else "Nein",
                    "Ja" if p.manage_roles else "Nein",
                    "Ja" if p.manage_guild else "Nein",
                    "Ja" if p.manage_webhooks else "Nein",
                    str(member.display_avatar.url) if member.display_avatar else "",
                    "Ja" if member.premium_since else "Nein",
                    fmt_dt(member.premium_since),
                    fmt_dt(member.timed_out_until) if member.timed_out_until and member.timed_out_until > datetime.now(timezone.utc) else "",
                    mc_username, "☐"
                ]
                row_vals = [all_values[i] for i in selected_indices]
                for col, val in enumerate(row_vals, 1):
                    c = ws1.cell(row=idx, column=col, value=val)
                    c.font = Font(name="Arial", size=10)
                    c.alignment = LEFT
                try:
                    col_hex_idx = selected_indices.index(15)
                    if role_color:
                        color_cell(ws1.cell(row=idx, column=col_hex_idx+1), f"{role_color:06X}")
                except ValueError: pass
                try:
                    col_mc_idx = selected_indices.index(16)
                    color_cell(ws1.cell(row=idx, column=col_mc_idx+1), mc_hex, bold=True)
                except ValueError: pass
                try:
                    col_mchex_idx = selected_indices.index(17)
                    color_cell(ws1.cell(row=idx, column=col_mchex_idx+1), mc_hex)
                except ValueError: pass

                if (idx-1) % 100 == 0:
                    self.log(f"   Mitglied {idx-1}/{total}")
            self.log(f"   ✅ Mitglieder-Tabelle fertig ({total} Einträge)")

        # Rollen
        if self.options.get("export_roles", True):
            ws2 = wb.create_sheet("Rollen")
            H2 = ["Rollen-ID", "Name", "Position", "Farbe (HEX)", "Minecraft-Farbe", "MC-HEX", "Mitglieder",
                  "Erwähnbar?", "Angeheftet?", "Bot?", "Admin?", "Msg verwalten?", "Kick?", "Ban?", "Kanäle?",
                  "Rollen?", "Erstellt am"]
            W2 = [20, 28, 10, 14, 18, 12, 12, 12, 20, 20, 14, 20, 8, 8, 14, 14, 18]
            apply_header(ws2, H2, W2)
            ws2.freeze_panes = "A2"
            ws2.auto_filter.ref = f"A1:{get_column_letter(len(H2))}1"
            role_counts = {}
            for m in members:
                for r in m.roles:
                    role_counts[r.id] = role_counts.get(r.id, 0) + 1
            roles_sorted = [r for r in sorted(guild.roles, key=lambda r: -r.position) if r.name != "@everyone"]
            for ri, role in enumerate(roles_sorted, 2):
                mc_n, mc_h = closest_minecraft_color(role.color.value)
                p = role.permissions
                vals = [str(role.id), role.name, role.position,
                        f"#{role.color.value:06X}" if role.color.value else "Standard",
                        mc_n, mc_h, role_counts.get(role.id, 0),
                        "Ja" if role.mentionable else "Nein", "Ja" if role.hoist else "Nein",
                        "Ja" if role.managed else "Nein", "Ja" if p.administrator else "Nein",
                        "Ja" if p.manage_messages else "Nein", "Ja" if p.kick_members else "Nein",
                        "Ja" if p.ban_members else "Nein", "Ja" if p.manage_channels else "Nein",
                        "Ja" if p.manage_roles else "Nein", fmt_dt(role.created_at)]
                for col, val in enumerate(vals, 1):
                    c = ws2.cell(row=ri, column=col, value=val)
                    c.font = Font(name="Arial", size=10)
                    c.alignment = LEFT
                if role.color.value:
                    color_cell(ws2.cell(row=ri, column=4), f"{role.color.value:06X}")
                color_cell(ws2.cell(row=ri, column=5), mc_h, bold=True)
                color_cell(ws2.cell(row=ri, column=6), mc_h)
            self.log(f"   ✅ Rollen-Tabelle fertig ({len(roles_sorted)} Rollen)")

        # Server-Info
        if self.options.get("export_serverinfo", True):
            ws3 = wb.create_sheet("Server-Info")
            ws3.column_dimensions["A"].width = 32
            ws3.column_dimensions["B"].width = 48
            ws3.merge_cells("A1:B1")
            t = ws3["A1"]
            t.value = f"Server-Info: {guild.name}"
            t.font = Font(bold=True, size=14, color="FFFFFF")
            t.fill = PatternFill("solid", start_color="5865F2")
            t.alignment = Alignment(horizontal="center", vertical="center")
            ws3.row_dimensions[1].height = 30
            online = sum(1 for m in members if m.status == discord.Status.online)
            idle = sum(1 for m in members if m.status == discord.Status.idle)
            dnd = sum(1 for m in members if m.status == discord.Status.dnd)
            offline = sum(1 for m in members if m.status == discord.Status.offline)
            boosters = sum(1 for m in members if m.premium_since)
            text_channels = len(guild.text_channels)
            voice_channels = len(guild.voice_channels)
            categories = len(guild.categories)
            forum_channels = stage_channels = 0
            for ch in guild.channels:
                if hasattr(ch, 'type'):
                    if ch.type == discord.ChannelType.forum: forum_channels += 1
                    elif ch.type == discord.ChannelType.stage_voice: stage_channels += 1
            boost_count = getattr(guild, 'premium_subscription_count', 0) or 0
            info = [
                ("── Allgemein ──", ""), ("Server-ID", str(guild.id)), ("Servername", guild.name),
                ("Beschreibung", guild.description or "–"), ("Erstellt am", fmt_dt(guild.created_at)),
                ("Server-Alter (Tage)", days_since(guild.created_at)),
                ("Besitzer", str(guild.owner) if guild.owner else str(guild.owner_id)),
                ("Vanity-URL", guild.vanity_url_code or "–"), ("Bevorzugte Sprache", str(guild.preferred_locale)),
                ("", ""), ("── Mitglieder ──", ""), ("Gesamt", guild.member_count), ("Menschen", humans),
                ("Bots", bots), ("Server-Booster", boosters), ("🟢 Online", online),
                ("🌙 Abwesend", idle), ("⛔ Bitte nicht stören", dnd), ("⚫ Offline", offline), ("", ""),
                ("── Kanäle ──", ""), ("Textkanäle", text_channels), ("Sprachkanäle", voice_channels),
                ("Kategorien", categories), ("Forum-Kanäle", forum_channels), ("Stage-Kanäle", stage_channels),
                ("Gesamt Kanäle", len(guild.channels)), ("", ""), ("── Rollen & Medien ──", ""),
                ("Anzahl Rollen", len(guild.roles)-1), ("Anzahl Emojis", len(guild.emojis)),
                ("Anzahl Sticker", len(guild.stickers)), ("", ""), ("── Server-Boost ──", ""),
                ("Boost-Stufe", guild.premium_tier), ("Anzahl Boosts", boost_count), ("", ""),
                ("── Sicherheit ──", ""), ("Verifizierungsstufe", str(guild.verification_level)),
                ("Explizite Inhalte Filter", str(guild.explicit_content_filter)),
                ("2FA für Moderatoren", "Ja" if guild.mfa_level.value else "Nein"), ("", ""),
                ("── Export-Metadaten ──", ""), ("Exportiert am", fmt_dt(datetime.now(timezone.utc))),
                ("Exportierte Mitglieder", len(members)), ("Exportierte Rollen", len(guild.roles)-1),
            ]
            KEY_F = Font(bold=True, size=10)
            VAL_F = Font(size=10)
            SEC_FILL = PatternFill("solid", start_color="23272A")
            for i, (key, val) in enumerate(info, 2):
                kc = ws3.cell(row=i, column=1, value=key)
                vc = ws3.cell(row=i, column=2, value=val)
                kc.alignment = LEFT; vc.alignment = LEFT
                if key.startswith("──"):
                    kc.font = Font(bold=True, color="FFFFFF"); vc.font = Font(color="FFFFFF")
                    kc.fill = SEC_FILL; vc.fill = SEC_FILL
                else:
                    kc.font = KEY_F; vc.font = VAL_F
            self.log("   ✅ Server-Info fertig")

        try:
            wb.save(self.output_path)
        except PermissionError:
            self.log(f"❌ Keine Schreibrechte für '{self.output_path}'")
        else:
            self.log(f"\n✅ Export abgeschlossen! Datei: {self.output_path}")
            self.log(f"   👥 {total} Mitglieder | 🤖 {bots} Bots | 👤 {humans} Menschen")
        finally:
            await self.close()

    def run_export(self):
        # Startet den Discord-Client (blockierend)
        try:
            self.run(self._token)
        except Exception as e:
            self.log(f"❌ Bot-Fehler: {e}")

# GUI (unverändert, nur Start-Methode ruft run_export auf)
class ExportGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Discord Members → Excel Exporter")
        self.root.geometry("700x800")
        self.root.resizable(True, True)
        self.token_var = tk.StringVar()
        self.guild_var = tk.StringVar()
        self.output_var = tk.StringVar(value="Mitglieder_Export.xlsx")
        self.export_members_var = tk.BooleanVar(value=True)
        self.export_roles_var = tk.BooleanVar(value=True)
        self.export_serverinfo_var = tk.BooleanVar(value=True)
        self.include_minecraft_name_var = tk.BooleanVar(value=True)
        self.column_vars = []
        self.column_names = [
            "User-ID", "Username", "Anzeigename (Server)", "Globaler Name", "Discriminator",
            "Bot?", "System-Account?", "Online-Status", "Account erstellt", "Server beigetreten",
            "Tage Mitglied", "Account-Alter (Tage)", "Anzahl Rollen", "Höchste Rolle",
            "Rollenfarbe (HEX)", "Minecraft-Farbe", "MC-Farbe (HEX)", "Alle Rollen",
            "Administrator?", "Nachrichten verwalten?", "Kick?", "Ban?", "Moderator (Timeout)?",
            "Kanäle verwalten?", "Rollen verwalten?", "Server verwalten?", "Webhooks verwalten?",
            "Avatar-URL", "Server-Booster?", "Boosted seit", "Timeout bis", "Minecraft-Name", "Gebaut"
        ]
        self.create_widgets()
        self.center_window()

    def center_window(self):
        self.root.update_idletasks()
        w, h = self.root.winfo_width(), self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (w // 2)
        y = (self.root.winfo_screenheight() // 2) - (h // 2)
        self.root.geometry(f"{w}x{h}+{x}+{y}")

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(main_frame, text="Bot-Token:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.token_var, width=60, show="*").grid(row=0, column=1, columnspan=2, sticky=tk.W, pady=2)
        ttk.Label(main_frame, text="Server-ID:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.guild_var, width=30).grid(row=1, column=1, sticky=tk.W, pady=2)
        ttk.Label(main_frame, text="Ausgabedatei:").grid(row=2, column=0, sticky=tk.W, pady=2)
        ttk.Entry(main_frame, textvariable=self.output_var, width=50).grid(row=2, column=1, sticky=tk.W, pady=2)
        ttk.Button(main_frame, text="Durchsuchen...", command=self.browse_output).grid(row=2, column=2, padx=5)
        ttk.Label(main_frame, text="Export-Optionen:", font=("Arial", 10, "bold")).grid(row=3, column=0, sticky=tk.W, pady=(15,5))
        ttk.Checkbutton(main_frame, text="Mitglieder exportieren", variable=self.export_members_var,
                        command=self.toggle_column_frame).grid(row=4, column=0, sticky=tk.W, padx=20)
        ttk.Checkbutton(main_frame, text="Rollen exportieren", variable=self.export_roles_var).grid(row=5, column=0, sticky=tk.W, padx=20)
        ttk.Checkbutton(main_frame, text="Server-Info exportieren", variable=self.export_serverinfo_var).grid(row=6, column=0, sticky=tk.W, padx=20)
        self.mc_check = ttk.Checkbutton(main_frame, text="Minecraft-Namen aus Supabase einfügen",
                                        variable=self.include_minecraft_name_var)
        if SUPABASE_URL and SUPABASE_KEY:
            self.mc_check.grid(row=7, column=0, sticky=tk.W, padx=20, pady=(10,0))
        else:
            self.include_minecraft_name_var.set(False)
        self.column_frame = ttk.LabelFrame(main_frame, text="Mitglieder-Spalten (ankreuzen = exportieren)", padding="5")
        self.column_frame.grid(row=8, column=0, columnspan=3, sticky="ew", pady=10, padx=20)
        self.column_frame.columnconfigure(0, weight=1)
        self.column_checkboxes = []
        for i, name in enumerate(self.column_names):
            var = tk.BooleanVar(value=True)
            self.column_vars.append(var)
            cb = ttk.Checkbutton(self.column_frame, text=name, variable=var)
            row, col = i // 4, i % 4
            cb.grid(row=row, column=col, sticky=tk.W, padx=5, pady=1)
            self.column_checkboxes.append(cb)
        btn_frame = ttk.Frame(self.column_frame)
        btn_frame.grid(row=(len(self.column_names)//4)+1, column=0, columnspan=4, pady=5)
        ttk.Button(btn_frame, text="Alle auswählen", command=self.select_all_columns).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Keine auswählen", command=self.deselect_all_columns).pack(side=tk.LEFT, padx=5)
        ttk.Label(main_frame, text="Log:", font=("Arial", 10, "bold")).grid(row=9, column=0, sticky=tk.W, pady=(10,0))
        self.log_text = scrolledtext.ScrolledText(main_frame, height=12, state=tk.DISABLED, wrap=tk.WORD)
        self.log_text.grid(row=10, column=0, columnspan=3, sticky="nsew", pady=5)
        main_frame.rowconfigure(10, weight=1)
        main_frame.columnconfigure(1, weight=1)
        self.start_btn = ttk.Button(main_frame, text="Export starten", command=self.start_export)
        self.start_btn.grid(row=11, column=0, columnspan=3, pady=15)

    def browse_output(self):
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if fn: self.output_var.set(fn)

    def toggle_column_frame(self):
        self.column_frame.grid() if self.export_members_var.get() else self.column_frame.grid_remove()

    def select_all_columns(self):
        for v in self.column_vars: v.set(True)
    def deselect_all_columns(self):
        for v in self.column_vars: v.set(False)

    def log(self, msg):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def start_export(self):
        token = self.token_var.get().strip()
        gid_str = self.guild_var.get().strip()
        out = self.output_var.get().strip()
        if not token or not gid_str:
            messagebox.showerror("Fehler", "Token und Server-ID erforderlich.")
            return
        try: gid = int(gid_str)
        except: messagebox.showerror("Fehler", "Server-ID muss Zahl sein."); return
        if not out: out = "Mitglieder_Export.xlsx"; self.output_var.set(out)
        selected = [i for i, v in enumerate(self.column_vars) if v.get()]
        if self.export_members_var.get() and not selected:
            messagebox.showerror("Fehler", "Mindestens eine Mitglieder-Spalte wählen."); return
        options = {
            "export_members": self.export_members_var.get(),
            "export_roles": self.export_roles_var.get(),
            "export_serverinfo": self.export_serverinfo_var.get(),
            "member_columns": selected,
            "include_minecraft_name": self.include_minecraft_name_var.get()
        }
        self.start_btn.config(state=tk.DISABLED)
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        def run():
            exporter = Exporter(token, gid, out, options, self.log)
            try:
                exporter.run_export()
            except Exception as e:
                self.log(f"❌ Fehler: {e}")
            finally:
                self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL))
        threading.Thread(target=run, daemon=True).start()

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    ExportGUI().run()